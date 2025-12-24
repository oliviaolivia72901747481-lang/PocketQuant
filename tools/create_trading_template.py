"""
科技股实盘监控Excel模板生成器

生成包含以下工作表的Excel模板：
1. 交易记录 - 记录每笔买卖交易
2. 当前持仓 - 实时持仓监控
3. 绩效统计 - 月度/年度绩效汇总
4. 策略参数 - v11.4g策略参数配置
5. 使用说明 - 模板使用指南
6. 股票池 - 60只科技股配置
"""
import sys
sys.path.insert(0, '.')

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.chart import LineChart, Reference
from datetime import datetime


def create_trading_template():
    """创建实盘监控Excel模板 - v11.4g版本"""
    wb = Workbook()
    
    # 定义样式
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 新增样式
    title_font = Font(bold=True, size=14, color="1F4E79")
    section_font = Font(bold=True, size=12, color="2E75B6")
    highlight_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    new_feature_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 数字格式
    currency_format = '#,##0.00'
    percent_format = '0.00%'
    date_format = 'YYYY-MM-DD'
    
    # ========== 1. 交易记录表 ==========
    ws_trades = wb.active
    ws_trades.title = "交易记录"
    
    # 表头
    trade_headers = [
        "序号", "交易日期", "股票代码", "股票名称", "交易方向",
        "买入价格", "卖出价格", "股数", "买入金额", "卖出金额",
        "盈亏金额", "盈亏比例", "持仓天数", "卖出原因", "备注"
    ]
    
    for col, header in enumerate(trade_headers, 1):
        cell = ws_trades.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 设置列宽
    trade_widths = [6, 12, 10, 12, 10, 10, 10, 8, 12, 12, 12, 10, 10, 15, 20]
    for col, width in enumerate(trade_widths, 1):
        ws_trades.column_dimensions[get_column_letter(col)].width = width
    
    # 添加示例数据行（公式）
    for row in range(2, 52):  # 预留50行
        ws_trades.cell(row=row, column=1, value=row-1)  # 序号
        # 盈亏金额公式
        ws_trades.cell(row=row, column=11).value = f"=IF(J{row}<>\"\",J{row}-I{row},\"\")"
        # 盈亏比例公式
        ws_trades.cell(row=row, column=12).value = f"=IF(I{row}<>\"\",K{row}/I{row},\"\")"
        ws_trades.cell(row=row, column=12).number_format = percent_format
    
    # 冻结首行
    ws_trades.freeze_panes = "A2"
    
    # ========== 2. 当前持仓表 ==========
    ws_position = wb.create_sheet("当前持仓")
    
    position_headers = [
        "序号", "股票代码", "股票名称", "买入日期", "买入价格",
        "持仓股数", "买入金额", "当前价格", "当前市值", "浮动盈亏",
        "盈亏比例", "持仓天数", "最高价", "最高盈利", "止损价",
        "止盈价", "移动止盈触发", "状态", "备注"
    ]
    
    for col, header in enumerate(position_headers, 1):
        cell = ws_position.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 设置列宽
    position_widths = [6, 10, 12, 12, 10, 10, 12, 10, 12, 12, 10, 10, 10, 10, 10, 10, 12, 10, 15]
    for col, width in enumerate(position_widths, 1):
        ws_position.column_dimensions[get_column_letter(col)].width = width
    
    # 添加公式 - v11.4g参数
    for row in range(2, 12):  # 预留10行（最多5只持仓）
        ws_position.cell(row=row, column=1, value=row-1)
        # 买入金额 = 买入价格 * 股数
        ws_position.cell(row=row, column=7).value = f"=IF(E{row}<>\"\",E{row}*F{row},\"\")"
        # 当前市值 = 当前价格 * 股数
        ws_position.cell(row=row, column=9).value = f"=IF(H{row}<>\"\",H{row}*F{row},\"\")"
        # 浮动盈亏 = 当前市值 - 买入金额
        ws_position.cell(row=row, column=10).value = f"=IF(I{row}<>\"\",I{row}-G{row},\"\")"
        # 盈亏比例
        ws_position.cell(row=row, column=11).value = f"=IF(G{row}<>\"\",J{row}/G{row},\"\")"
        ws_position.cell(row=row, column=11).number_format = percent_format
        # 持仓天数
        ws_position.cell(row=row, column=12).value = f"=IF(D{row}<>\"\",TODAY()-D{row},\"\")"
        # 最高盈利
        ws_position.cell(row=row, column=14).value = f"=IF(M{row}<>\"\",(M{row}-E{row})/E{row},\"\")"
        ws_position.cell(row=row, column=14).number_format = percent_format
        # 止损价 = 买入价 * (1 - 4.6%) - v11.4g更新
        ws_position.cell(row=row, column=15).value = f"=IF(E{row}<>\"\",E{row}*0.954,\"\")"
        # 止盈价 = 买入价 * (1 + 22%) - v11.4g更新
        ws_position.cell(row=row, column=16).value = f"=IF(E{row}<>\"\",E{row}*1.22,\"\")"
        # 移动止盈触发价 = 买入价 * (1 + 9%)
        ws_position.cell(row=row, column=17).value = f"=IF(E{row}<>\"\",E{row}*1.09,\"\")"
    
    # 汇总行
    ws_position.cell(row=13, column=6, value="合计:")
    ws_position.cell(row=13, column=6).font = Font(bold=True)
    ws_position.cell(row=13, column=7).value = "=SUM(G2:G11)"
    ws_position.cell(row=13, column=9).value = "=SUM(I2:I11)"
    ws_position.cell(row=13, column=10).value = "=SUM(J2:J11)"
    ws_position.cell(row=13, column=11).value = "=IF(G13<>\"\",J13/G13,\"\")"
    ws_position.cell(row=13, column=11).number_format = percent_format
    
    ws_position.freeze_panes = "A2"
    
    # ========== 3. 绩效统计表 ==========
    ws_perf = wb.create_sheet("绩效统计")
    
    # 账户概览
    ws_perf.cell(row=1, column=1, value="账户概览").font = Font(bold=True, size=14)
    ws_perf.merge_cells('A1:D1')
    
    overview_data = [
        ["初始资金", 100000, "", ""],
        ["当前总资产", "=当前持仓!G13+B4", "", ""],
        ["可用现金", "", "", ""],
        ["持仓市值", "=当前持仓!I13", "", ""],
        ["累计盈亏", "=B3-B2", "", ""],
        ["累计收益率", "=B6/B2", "", ""],
        ["", "", "", ""],
        ["本月统计", "", "", ""],
        ["本月交易次数", "=COUNTIF(交易记录!B:B,\">=\"&DATE(YEAR(TODAY()),MONTH(TODAY()),1))", "", ""],
        ["本月盈利次数", "", "", ""],
        ["本月亏损次数", "", "", ""],
        ["本月胜率", "", "", ""],
        ["本月盈亏", "", "", ""],
    ]
    
    for row, data in enumerate(overview_data, 2):
        for col, value in enumerate(data, 1):
            cell = ws_perf.cell(row=row, column=col, value=value)
            if col == 1:
                cell.font = Font(bold=True)
    
    ws_perf.cell(row=7, column=2).number_format = percent_format
    
    # 月度绩效表
    ws_perf.cell(row=16, column=1, value="月度绩效").font = Font(bold=True, size=14)
    
    monthly_headers = ["月份", "交易次数", "盈利次数", "亏损次数", "胜率", "总盈亏", "收益率", "最大回撤"]
    for col, header in enumerate(monthly_headers, 1):
        cell = ws_perf.cell(row=17, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 预留12个月
    months = ["2025-01", "2025-02", "2025-03", "2025-04", "2025-05", "2025-06",
              "2025-07", "2025-08", "2025-09", "2025-10", "2025-11", "2025-12"]
    for row, month in enumerate(months, 18):
        ws_perf.cell(row=row, column=1, value=month)
    
    # ========== 4. 策略参数表 - v11.4g ==========
    ws_params = wb.create_sheet("策略参数")
    
    ws_params.cell(row=1, column=1, value="科技股策略 v11.4g 参数配置（平衡版）").font = Font(bold=True, size=14)
    ws_params.merge_cells('A1:D1')
    
    # 版本说明
    ws_params.cell(row=2, column=1, value="回测绩效：收益33.51% | 回撤-4.81% | 胜率24.5% | 收益/回撤比6.96")
    ws_params.cell(row=2, column=1).font = Font(italic=True, color="666666")
    ws_params.merge_cells('A2:D2')
    
    params_data = [
        ["", "", "", ""],
        ["仓位管理", "", "", ""],
        ["单只股票仓位", "11%", "控制单笔风险", ""],
        ["最大持仓数量", "5只", "分散风险", ""],
        ["", "", "", ""],
        ["止盈止损", "", "", ""],
        ["止损线", "-4.6%", "严格止损，控制亏损", "v11.4g调整"],
        ["止盈线", "+22%", "锁定利润", "v11.4g调整"],
        ["移动止盈触发", "+9%", "盈利9%后启动移动止盈", ""],
        ["移动止盈回撤", "2.8%", "从最高点回撤2.8%卖出", ""],
        ["最大持仓天数", "15天", "避免长期套牢", ""],
        ["", "", "", ""],
        ["买入条件", "", "", ""],
        ["MA金叉", "MA5上穿MA20", "趋势确认", ""],
        ["中期趋势", "股价>MA60", "中期向上", ""],
        ["RSI范围", "44~70", "避免超买超卖", "v11.4g调整"],
        ["MACD确认", "柱状图向上或为正", "动能确认", ""],
        ["成交量确认", "当日量>5日均量×1.1", "放量确认", "v11.4g调整"],
        ["信号强度门槛", "≥83", "只选择高质量信号", "v11.4g新增"],
        ["趋势过滤", "MA20斜率向上", "确保上升趋势", "v11.4g新增"],
        ["价格位置", "价格<MA5×1.05", "避免追高", "v11.4g新增"],
        ["", "", "", ""],
        ["卖出条件", "", "", ""],
        ["止损", "亏损≥4.6%", "严格止损", "v11.4g调整"],
        ["止盈", "盈利≥22%", "锁定利润", "v11.4g调整"],
        ["移动止盈", "盈利曾达9%，回撤2.8%", "保护利润", ""],
        ["RSI超买", "RSI>80且盈利", "仅盈利时触发", "v11.4g调整"],
        ["趋势反转", "MA5<MA20且亏损", "趋势破坏", ""],
        ["持仓超时", "持仓≥15天", "避免套牢", ""],
    ]
    
    for row, data in enumerate(params_data, 3):
        for col, value in enumerate(data, 1):
            cell = ws_params.cell(row=row, column=col, value=value)
            if col == 1 and value and value not in ["仓位管理", "止盈止损", "买入条件", "卖出条件"]:
                pass
            elif value in ["仓位管理", "止盈止损", "买入条件", "卖出条件"]:
                cell.font = section_font
            # 高亮v11.4g新增/调整项
            if col == 4 and value and ("新增" in value or "调整" in value):
                for c in range(1, 5):
                    ws_params.cell(row=row+3, column=c).fill = new_feature_fill
    
    # 设置列宽
    ws_params.column_dimensions['A'].width = 18
    ws_params.column_dimensions['B'].width = 25
    ws_params.column_dimensions['C'].width = 25
    ws_params.column_dimensions['D'].width = 15
    
    # ========== 5. 使用说明表 ==========
    ws_guide = wb.create_sheet("使用说明")
    
    ws_guide.cell(row=1, column=1, value="科技股实盘监控模板使用说明 v11.4g").font = Font(bold=True, size=16)
    ws_guide.merge_cells('A1:D1')
    
    guide_content = [
        "",
        "一、交易记录表",
        "   1. 每次买入/卖出后，在此表记录交易详情",
        "   2. 盈亏金额和盈亏比例会自动计算",
        "   3. 卖出原因请填写：止损/止盈/移动止盈/RSI超买/趋势反转/持仓超时",
        "",
        "二、当前持仓表",
        "   1. 买入后在此表添加持仓记录",
        "   2. 每日更新'当前价格'和'最高价'",
        "   3. 止损价(-4.6%)、止盈价(+22%)、移动止盈触发价(+9%)会自动计算",
        "   4. 当价格触及止损/止盈价时，执行卖出",
        "   5. 当最高盈利≥9%且当前盈利回撤2.8%时，执行移动止盈",
        "",
        "三、绩效统计表",
        "   1. 账户概览显示总体绩效",
        "   2. 月度绩效需手动汇总填写",
        "   3. 建议每月末进行绩效复盘",
        "",
        "四、策略参数表",
        "   1. 记录v11.4g策略的所有参数",
        "   2. 严格按照参数执行，不要主观判断",
        "   3. 绿色高亮项为v11.4g版本新增/调整的参数",
        "",
        "五、v11.4g版本核心改进",
        "   1. 止损从-4.5%调整为-4.6%（略放宽）",
        "   2. 止盈从+20%调整为+22%（提高目标）",
        "   3. RSI范围从55-80调整为44-70（更宽松）",
        "   4. 量比从1.5调整为1.1（降低门槛）",
        "   5. 新增信号强度门槛≥83（只选高质量信号）",
        "   6. 新增趋势过滤（MA20斜率向上）",
        "   7. 新增价格位置过滤（避免追高）",
        "   8. RSI超买卖出仅在盈利时触发",
        "",
        "六、操作流程",
        "   1. 每日开盘前：检查系统买入信号",
        "   2. 买入后：在'当前持仓'表添加记录",
        "   3. 每日收盘后：更新当前价格和最高价",
        "   4. 触发卖出条件时：执行卖出，记录到'交易记录'表",
        "   5. 每月末：汇总月度绩效",
        "",
        "七、风险提示",
        "   1. 严格执行止损，不要抱有侥幸心理",
        "   2. 单只股票仓位不超过11%",
        "   3. 最多同时持有5只股票",
        "   4. 股市有风险，投资需谨慎",
        "",
        "八、v11.4g回测绩效参考",
        "   - 回测周期：2022-12-26 ~ 2024-12-20（约2年）",
        "   - 总收益率：33.51%",
        "   - 最大回撤：-4.81%（远低于15%警戒线）",
        "   - 胜率：24.5%",
        "   - 收益/回撤比：6.96（优秀）",
        "   - 月均交易：11.6次",
    ]
    
    for row, content in enumerate(guide_content, 3):
        ws_guide.cell(row=row, column=1, value=content)
    
    ws_guide.column_dimensions['A'].width = 80
    
    # ========== 6. 股票池表 ==========
    ws_pool = wb.create_sheet("股票池")
    
    ws_pool.cell(row=1, column=1, value="科技股股票池（60只）").font = Font(bold=True, size=14)
    ws_pool.merge_cells('A1:D1')
    
    pool_headers = ["序号", "股票代码", "股票名称", "所属行业"]
    for col, header in enumerate(pool_headers, 1):
        cell = ws_pool.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 导入股票池数据
    try:
        from config.tech_stock_pool import TECH_STOCK_POOL
        row = 3
        for sector, stocks in TECH_STOCK_POOL.items():
            for stock in stocks:
                ws_pool.cell(row=row, column=1, value=row-2)
                ws_pool.cell(row=row, column=2, value=stock["code"])
                ws_pool.cell(row=row, column=3, value=stock["name"])
                ws_pool.cell(row=row, column=4, value=sector)
                row += 1
    except Exception as e:
        print(f"警告：无法导入股票池数据: {e}")
    
    # 设置列宽
    ws_pool.column_dimensions['A'].width = 8
    ws_pool.column_dimensions['B'].width = 12
    ws_pool.column_dimensions['C'].width = 15
    ws_pool.column_dimensions['D'].width = 15
    
    # ========== 7. 参数速查卡 ==========
    ws_card = wb.create_sheet("参数速查卡")
    
    ws_card.cell(row=1, column=1, value="科技股策略 v11.4g 参数速查卡").font = Font(bold=True, size=16)
    ws_card.merge_cells('A1:C1')
    
    card_data = [
        ["", "", ""],
        ["【止损止盈】", "", ""],
        ["止损线", "-4.6%", ""],
        ["止盈线", "+22%", ""],
        ["移动止盈触发", "+9%", ""],
        ["移动止盈回撤", "2.8%", ""],
        ["最大持仓天数", "15天", ""],
        ["", "", ""],
        ["【仓位管理】", "", ""],
        ["单只仓位", "11%", ""],
        ["最大持仓", "5只", ""],
        ["", "", ""],
        ["【买入条件】", "", ""],
        ["MA金叉", "MA5>MA20", ""],
        ["中期趋势", "价格>MA60", ""],
        ["RSI范围", "44~70", ""],
        ["量比", ">1.1", ""],
        ["信号强度", "≥83", ""],
        ["趋势过滤", "MA20斜率↑", ""],
        ["价格位置", "<MA5×1.05", ""],
        ["", "", ""],
        ["【卖出条件】", "", ""],
        ["止损", "亏损≥4.6%", ""],
        ["止盈", "盈利≥22%", ""],
        ["移动止盈", "曾达9%回撤2.8%", ""],
        ["RSI超买", "RSI>80且盈利", ""],
        ["趋势反转", "MA5<MA20且亏损", ""],
        ["持仓超时", "≥15天", ""],
    ]
    
    for row, data in enumerate(card_data, 2):
        for col, value in enumerate(data, 1):
            cell = ws_card.cell(row=row, column=col, value=value)
            if value and value.startswith("【"):
                cell.font = section_font
                cell.fill = highlight_fill
    
    ws_card.column_dimensions['A'].width = 18
    ws_card.column_dimensions['B'].width = 20
    ws_card.column_dimensions['C'].width = 15
    
    # 保存文件
    filename = f"科技股实盘监控模板_v11.4g_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"✅ 实盘监控Excel模板已生成: {filename}")
    
    return filename


if __name__ == "__main__":
    create_trading_template()
